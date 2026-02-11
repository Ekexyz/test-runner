import time
from robot.api import logger
from openpyxl import load_workbook

class ExcelHelper:
    """Excel operations with performance timing"""
    
    def read_excel_data(self, file_path, sheet_name, header_row=1, data_row=2, max_rows=100):
        """Read Excel data using openpyxl with bulk reading"""
        start_time = time.time()
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[sheet_name]
        
        # Read headers
        headers = [cell.value for cell in sheet[header_row]]
        # logger.info(f"Headers read: {headers}")
        
        # Read all rows at once
        rows = []
        actual_row = data_row
        
        for row in sheet.iter_rows(min_row=data_row, max_row=data_row + max_rows - 1, values_only=True):
            if row[0] is None:
                break
            row_dict = [{"key": headers[i], "value": row[i] if i < len(row) else None} for i in range(len(headers))]
            rows.append(row_dict)
            actual_row += 1
        
        wb.close()
        
        end_time = time.time()
        elapsed = end_time - start_time
        
        logger.console(f"⏱️ [PYTHON-OPENPYXL] Read {len(rows)} rows in {elapsed:.4f} seconds ({elapsed*1000:.2f} ms)")
        logger.console(f"⏱️ [PYTHON-OPENPYXL] Average per row: {(elapsed/len(rows))*1000:.2f} ms" if rows else "No rows read")
        
        return rows